import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from common import stats

_CHARS_PARAMS = ["角色名称", "基础攻击", "攻击力", "暴击率", "暴击伤害", "常驻增伤"]


def output_worksheet(ws, dmg_data):
    '''
    :param ws: excel worksheet
    :param dmg_data: {enemy_name: {team_name:{char_name: dmg}}}
    :return:
    '''
    ws.merge_cells('A1:A2')
    ws.merge_cells('B1:B2')
    # .alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(1, 1).value = "队伍"
    ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions[get_column_letter(1)].width = 14
    ws.cell(1, 2).value = "角色"
    ws.cell(1, 2).alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions[get_column_letter(2)].width = 23
    for i, enemy_name in enumerate(dmg_data):
        # print(enemy_name)
        ws.cell(1, 3+i*2).value = enemy_name
        ws.cell(1, 3+i*2).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(2, 3+i*2).value = "伤害数值"
        ws.cell(2, 3+i*2).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(2, 4+i*2).value = "伤害占比"
        ws.cell(2, 4+i*2).alignment = Alignment(horizontal="center", vertical="center")

        ws.column_dimensions[get_column_letter(3+i*2)].width = 18

        ws.merge_cells(start_row=1, start_column=3+i*2, end_row=1, end_column=4+i*2)
        row_count = 2
        # print(dmg_data[enemy_name])
        for j, team_name in enumerate(dmg_data[enemy_name]):
            # print(team_name)
            chars_dmg = dmg_data[enemy_name][team_name]
            # print(chars_dmg)
            ws.merge_cells(start_row=row_count+1, start_column=1, end_row=row_count+len(chars_dmg), end_column=1)
            ws.cell(row_count + 1, 1).value = team_name
            ws.cell(row_count + 1, 1).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row_count + 1, 1).font = Font(bold=True)

            # print(chars_dmg.keys())
            total_dmg = chars_dmg["总伤"]
            for k, char_name in enumerate(chars_dmg):
                ws.cell(row_count + k + 1, 2).value = char_name
                ws.cell(row_count + k + 1, 2).alignment = Alignment(horizontal="center", vertical="center")

                ws.cell(row_count + k + 1, 3+i*2).value = chars_dmg[char_name]  # "{:.4f}".format()
                ws.cell(row_count + k + 1, 3+i*2).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row_count + k + 1, 3+i*2).number_format = '0.00'

                ws.cell(row_count + k + 1, 4+i*2).value = chars_dmg[char_name]/total_dmg  # "{:.4f}".format()
                ws.cell(row_count + k + 1, 4+i*2).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row_count + k + 1, 4+i*2).number_format = '0.00%'

            row_count += k + 1
            ws.cell(row_count, 2).font = Font(bold=True)
            ws.cell(row_count, 3+i*2).font = Font(bold=True)


def save_stats(data, direc):
    wb = Workbook()
    ws = wb.active
    for row, char in enumerate(data):
        for col, value in enumerate(char.get_info()):
            if isinstance(value, stats.Character):
                ws.cell(row+1, col+1).value = value.name
            else:
                ws.cell(row+1, col+1).value = value
    # print(chars)
    wb.save(direc)
    # return chars


def load_chars(direc=r".\data\characters.xlsx"):
    ws = openpyxl.load_workbook(direc, data_only=True).worksheets[0]
    chars = []
    for row in ws.iter_rows(min_row=1, max_col=6):
        char = stats.Character(*(cell.value for cell in row))
        chars.append(char)
    # print(chars)
    return chars


def load_enemies(direc=r".\data\enemy.xlsx"):
    ws = openpyxl.load_workbook(direc, data_only=True).worksheets[0]
    enemies = []
    for row in ws.iter_rows(min_row=1, max_col=11):
        enemy = stats.Enemy(*(cell.value for cell in row))
        enemies.append(enemy)
    # print(chars)
    return enemies


def char_dict(chars):
    return {char.name: char for char in chars}


def load_skills(chars, direc=r".\data\skills.xlsx"):
    wb = openpyxl.load_workbook(direc, data_only=True)
    teams = []
    iterator = iter(wb.sheetnames)
    for sheet in wb.worksheets:
        name = next(iterator)
        skills = []
        remark = sheet['A1'].value
        sheet = sheet[2:sheet.max_row]
        for row in sheet:
            info = [cell.value for cell in row]
            info[1] = chars[info[1]]
            skill = stats.Skill(*info)
            skills.append(skill)
        team = stats.Team(name, remark, *skills)
        teams.append(team)
    return teams
