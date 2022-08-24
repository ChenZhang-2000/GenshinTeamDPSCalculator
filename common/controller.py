import re

import yaml
import openpyxl
from openpyxl.cell.cell import MergedCell
import torch
import pandas as pd

from common.stats import STATS_LENGTH, Infusion
from common.model import Model, Team
from common.characters import CHAR_FACTORY
from common.weapon import WEAPON_FACTORY
from common.artifact import ArtifactSet, ARTIFACT_FACTORY
from common.enemy import ENEMY_FACTORY
from common import config


def team_generation(char_data):
    chars = []
    for data in char_data:
        artifact_set = ArtifactSet(data['artifact']['stats'],
                                   two_set=ARTIFACT_FACTORY[data['artifact']['two_set']],
                                   four_set=ARTIFACT_FACTORY[data['artifact']['four_set']])
        weapon = WEAPON_FACTORY[data['weapon']['name']](data['weapon']['affix'])
        character = CHAR_FACTORY[data['char']['name']](weapon=weapon,
                                                       artifact=artifact_set,
                                                       level=data['char']['level'],
                                                       constellation=data['char']['constellation'],
                                                       name=data['name'],
                                                       ascension_phase=data['char']['ascension_phase'],
                                                       skill_level=data['char']['skill_level'])
        chars.append(character)
    team = Team(*chars)
    return team


def read_char_excel(file_direc=r".\data\characters.xlsx"):
    ws = openpyxl.load_workbook(file_direc, data_only=True).worksheets[0]
    chars = []
    header = [config.stats_map[cell.value] for cell in ws[1]]
    data_map = config.stats_pos_map
    for row in ws.iter_rows(min_row=2):
        artifact_data = torch.zeros(STATS_LENGTH)
        char_data = {}
        for i, cell in enumerate(row):
            value = cell.value
            if i <= 8:
                char_data[header[i]] = value
            else:
                value = 0. if value is None else float(value)
                artifact_data[data_map[header[i]]] = value
        level = char_data['character_level']
        if re.match(r"\d{1,}\+", level):
            level = int(level[:-1])
            ascension = (level >= torch.tensor([20, 40, 50, 60, 70, 80])).sum().item()
        else:
            level = int(level)
            ascension = (level > torch.tensor([20, 40, 50, 60, 70, 80])).sum().item()

        char_data = {'name': char_data['name'],
                     'char': {'name': config.char_map[char_data['character']],
                              'level': level,
                              'constellation': int(char_data['character_constellation']),
                              'ascension_phase': ascension,
                              'skill_level': tuple(int(i.strip()) for i in char_data['skill_level'].split(','))},
                     'weapon': {'name': config.weapon_map[char_data['weapon']],
                                'affix': int(char_data['weapon_affix'])},
                     'artifact': {'two_set': config.artifact_map[char_data['artifact_two']],
                                  'four_set': config.artifact_map[char_data['artifact_four']],
                                  'stats': artifact_data.reshape(1, STATS_LENGTH)}}
        # print(char_data['char']['skill_level'])
        chars.append(char_data)
    return chars


def read_enemy_excel(file_direc=r".\data\enemy.xlsx"):
    ws = openpyxl.load_workbook(file_direc, data_only=True).worksheets[0]
    enemies = {}
    header = [config.enemy_header_map[cell.value] for cell in ws[1]]
    for row in ws.iter_rows(min_row=2):
        enemy_data = {}
        for i, cell in enumerate(row):
            value = cell.value
            enemy_data[header[i]] = value
        enemies[enemy_data['name']] = ENEMY_FACTORY[config.enemy_map[enemy_data['enemy']]](enemy_data['enemy_level'])
    return enemies


def expand_params(params):
    arg = []
    kwarg = {}
    for param in params:
        if "=" in param:
            k, v = param.split("=")
            kwarg[k] = v
        else:
            arg.append(param)
    return arg, kwarg


def value_parsing(char, value, mode):
    objects = []
    if mode in ['skills', 'skill']:
        d = char.skills
    elif mode in ['buffs', 'buff']:
        d = char.buffs
    else:
        raise ValueError
    # print(value)
    for groups in re.findall(r"([a-zA-z\u4e00-\u9fff]{1,})(?:{{0,1})(\w{0,})(?:}{0,1})", value):
        name, params = groups
        params = [i.strip() for i in params.split(',')]
        if name in config.skill_map.keys():
            if config.skill_map[name] == 'weapon':
                if mode in ['skills', 'skill']:
                    target = char.weapon.skills
                elif mode in ['buffs', 'buff']:
                    target = char.weapon.buffs
                else:
                    raise ValueError

                if len(params) == 1 and params[0] == '':
                    objects.append(target[0])
                else:
                    arg, kwarg = expand_params(params[1:])
                    obj = target[int(params[0])]
                    obj.update(*arg, **kwarg)
                    objects.append(obj)

            elif config.skill_map[name] == 'artifact':
                if mode in ['skills', 'skill']:
                    target = char.artifact.skills
                elif mode in ['buffs', 'buff']:
                    target = char.artifact.buffs
                else:
                    raise ValueError

                if len(params) == 1 and params[0] == '':
                    objects.append(target[0])
                else:
                    arg, kwarg = expand_params(params[1:])
                    obj = target[int(params[0])]
                    obj.update(*arg, **kwarg)
                    objects.append(obj)
        else:
            arg, kwarg = expand_params(params)
            # if value == "附魔":
            #     print(d[name])
            obj = d[name]
            obj.update(*arg, **kwarg)
            objects.append(obj)
    return objects


def read_skill_excel(team, file_direc=r".\data\skills.xlsx"):
    char_map = {char.name: char for char in team}
    skill_df = pd.DataFrame(columns=['skill', 'start_time', 'end_time', 'reaction', 'kwarg'])
    buff_df = pd.DataFrame(columns=['buff', 'start_time', 'end_time'])
    infusion_df = pd.DataFrame(columns=['infusion', 'start_time', 'end_time'])

    ws = openpyxl.load_workbook(file_direc, data_only=True).worksheets[0]

    time_col, on_field_col, skill_col, buff_col = None, None, None, None
    col_char_map = {}
    for i, cell in enumerate(ws['A'], 1):
        try:
            mark = config.skill_header_map[cell.value]
            if mark == 'time':
                time_col = i
            elif mark == 'on_field':
                on_field_col = i
            elif mark == 'skill':
                skill_col = i
            elif mark == 'buff':
                buff_col = i
            else:
                raise
        except KeyError:
            # print(i)
            col_char_map[i] = char_map[cell.value]

    times = [float(cell.value) for cell in ws[time_col][1:]]
    # on_fields =
    # print(col_char_map)
    # print(buff_col)
    for idx, row in enumerate(ws.iter_rows(min_row=skill_col+1, max_row=buff_col-1), skill_col+1):
        # print(idx)
        char = col_char_map[idx]
        start = False
        i = 0
        for i, cell in enumerate(row[1:]):
            if i == len(times):
                break
            if isinstance(cell, MergedCell):
                pass
            else:
                if cell.value is None:
                    pass
                else:
                    if start:
                        end_time = times[i]
                        for skill in skills:
                            skill_df.loc[len(skill_df.index)] = [skill, start_time, end_time, None, {}]

                    start_time = times[i]
                    skills = value_parsing(char, cell.value, 'skills')
                    if len(skills) != 0:
                        start = True
                    else:
                        start = False
        if start:
            end_time = times[i]
            for skill in skills:
                skill_df.loc[len(skill_df.index)] = [skill, start_time, end_time, None, {}]

    for idx, row in enumerate(ws.iter_rows(min_row=buff_col + 1), buff_col + 1):
        # print(idx)
        char = col_char_map[idx]
        start = False
        i = 0
        for i, cell in enumerate(row[1:]):
            if i == len(times):
                pass
            if isinstance(cell, MergedCell):
                pass
            else:
                if cell.value is None:
                    pass
                else:
                    if start:
                        end_time = times[i]
                        for buff in buffs:
                            if isinstance(buff, Infusion):
                                infusion_df.loc[len(infusion_df.index)] = [buff, start_time, end_time]
                            else:
                                buff_df.loc[len(buff_df.index)] = [buff, start_time, end_time]

                    start_time = times[i]
                    buffs = value_parsing(char, cell.value, 'buffs')
                    if len(buffs) != 0:
                        start = True
                    else:
                        start = False
        if start:
            end_time = times[i]
            for buff in buffs:
                if isinstance(buff, Infusion):
                    infusion_df.loc[len(infusion_df.index)] = [buff, start_time, end_time]
                else:
                    buff_df.loc[len(buff_df.index)] = [buff, start_time, end_time]

    return skill_df, buff_df, infusion_df
